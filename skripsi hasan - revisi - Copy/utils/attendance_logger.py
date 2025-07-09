from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.protection import SheetProtection
from datetime import datetime, timedelta
import os

EXCEL_FILE = "data/attendance.xlsx"

def log_attendance(name):
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    # Buat file jika belum ada
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(["Name", "Date", "Time"])

        # Proteksi sheet agar tidak bisa diedit
        ws.protection.sheet = True
        ws.protection.enable()
        ws.protection.password = "absensi123"  # Password untuk proteksi sheet

        wb.save(EXCEL_FILE)

    # Baca file Excel yang sudah ada
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Attendance"]

    # Cek apakah orang ini sudah absen hari ini
    for row in reversed(list(ws.iter_rows(min_row=2, values_only=True))):  # baca dari bawah (paling baru)
        existing_name, existing_date, existing_time = row

        if existing_name == name and existing_date == date_str:
            last_absen_time = datetime.strptime(f"{existing_date} {existing_time}", "%Y-%m-%d %H:%M:%S")
            time_diff = now - last_absen_time

            if time_diff < timedelta(hours=1):
                print(f"[INFO] {name} sudah absen pada {last_absen_time}, tunggu {int((timedelta(hours=1) - time_diff).total_seconds() // 60)} menit lagi.")
                wb.close()
                return  # Tidak catat absen lagi

            break  # Sudah ketemu entri terakhir, lanjut catat

    # Kalau belum pernah absen hari ini atau sudah lewat 1 jam
    ws.protection.sheet = False  # Nonaktifkan proteksi sementara
    ws.append([name, date_str, time_str])
    ws.protection.sheet = True   # Aktifkan kembali proteksi sheet
    ws.protection.password = "absensi123"  # Pastikan password tetap ada

    wb.save(EXCEL_FILE)
    print(f"[INFO] Absensi {name} dicatat pada {date_str} {time_str}")
